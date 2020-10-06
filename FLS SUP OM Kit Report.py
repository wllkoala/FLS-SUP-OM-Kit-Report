from tkinter import Tk, filedialog

import pandas as pd
from openpyxl import load_workbook

root = Tk()
root.withdraw()
# 打开FLS SUP OM Kit Report
om_kit_report = filedialog.askopenfilename(
    title="FLS SUP OM Kit Report", filetypes=[("All files", "*")]
)
# 生成原始df
df_om = pd.read_excel(om_kit_report, sheet_name="FLS SUP OM Kit Report", skiprows=1)
# 生成判断bare pump item的df
df_barepump = pd.read_excel(om_kit_report, sheet_name="ITEM", usecols="A")
# 去除空项
df_barepump = df_barepump.dropna(axis=0, how="any")
# 将df内容用‘|’进行连接
sort_barepump = []
sort_barepump.append("|".join([str(item) for item in df_barepump.iloc[:, 0]]))
# 生成判断complete unit的df
df_completeunit = pd.read_excel(om_kit_report, sheet_name="ITEM", usecols="B")
# 去除空项
df_completeunit = df_completeunit.dropna(axis=0, how="any")
# 将df内容用‘|’进行连接
sort_completeunit = []
sort_completeunit.append("|".join([str(item) for item in df_completeunit.iloc[:, 0]]))
# 打开Sales order data report
sales_order_data_report = filedialog.askopenfilename(
    title="FLS CN Sales Order Data Report by Creation Date",
    filetypes=[("All files", "*")],
)
df_so = pd.read_excel(
    sales_order_data_report, sheet_name="FLS CN Sales Order Data Report "
)
# 原始df_om去除重复的End Item
df2_om = df_om.drop_duplicates(subset=["End Item"])
# 提取部分列
df2_om = df2_om[
    [
        "Project Num",
        "Supplymake Planner Code",
        "End So Num",
        "Customer Name",
        "End So Line",
        "End So Scheduled Date",
        "End Item Description",
        "End Item",
    ]
]
# 提取部分列
df2_so = df_so[
    [
        "Order Number",
        "Line Number",
        "Item Number",
        "Product Type",
        "Pump Size",
        "Ordered Quantity",
    ]
]
# 按照OM Kit对列重命名
df2_so = df2_so.rename(
    columns={
        "Order Number": "End So Num",
        "Line Number": "End So Line",
        "Item Number": "End Item",
        "Product Type": "Category",
        "Pump Size": "Pump Family",
    }
)
# 去除重复行
df2_so = df2_so.drop_duplicates(
    subset=["End So Num", "End So Line", "End Item", "Category", "Pump Family"]
)
# 将df按照关键列合并
df_summary = pd.merge(
    df2_om, df2_so, on=["End So Num", "End So Line", "End Item"], how="inner"
)
# 插入列‘光泵齐套情况’、‘整泵齐套情况’
df_summary.insert(df_summary.shape[1], "光泵齐套情况", "")
df_summary.insert(df_summary.shape[1], "整泵齐套情况", "")
for i in range(len(df_summary)):
    df1 = df_om[df_om["End Item"].isin([df_summary.loc[i, "End Item"]])]
    index_list = list(df1[df1["Supplymake Item"].str.contains(sort_barepump[0])].index)
    material_status = []
    for n in index_list:
        if not str(df_om.loc[n, "物料齐套"]) in material_status:
            material_status.append(str(df_om.loc[n, "物料齐套"]))
    material_status.sort()
    material_status = "|".join(material_status)
    material_status = material_status.replace(" 00:00:00", "")
    df_summary.loc[i, "光泵齐套情况"] = material_status
for i in range(len(df_summary)):
    df1 = df_om[df_om["End Item"].isin([df_summary.loc[i, "End Item"]])]
    index_list = list(
        df1[df1["Supplymake Item"].str.contains(sort_completeunit[0])].index
    )
    material_status = []
    for n in index_list:
        if not str(df_om.loc[n, "物料齐套"]) in material_status:
            material_status.append(str(df_om.loc[n, "物料齐套"]))
    material_status.sort()
    material_status = "|".join(material_status)
    material_status = material_status.replace(" 00:00:00", "")
    df_summary.loc[i, "整泵齐套情况"] = material_status
# 装载excel
wb = load_workbook(om_kit_report)
# 判断是否存在sheet名为Summary，如果有则删除
if "Summary" in wb.sheetnames:
    wb.remove(wb["Summary"])
# 如果有多个模块可以读写excel文件，这里要指定engine，否则可能会报错
writer = pd.ExcelWriter(om_kit_report, engine="openpyxl")
# 没有下面这个语句的话excel表将完全被覆盖
writer.book = wb
# 将df_summary写入Excel
df_summary.to_excel(writer, sheet_name="Summary", index=None)
writer.save()
writer.close()
